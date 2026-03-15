"""
NLDC XLS Parser
===============
Reads the TimeSeries sheet from NLDC daily PSP .xls/.xlsx files
and inserts all 96 rows into Neon Postgres.

Usage
-----
Single file:
    python parse_xls.py --file data/12.03.26_NLDC_PSP_568.xls

Whole folder:
    python parse_xls.py --folder data/

Setup DB (run once):
    python parse_xls.py --setup
"""

import os
import re
import sys
import argparse
from pathlib import Path
from datetime import date, datetime

import psycopg2
import xlrd           # old .xls (Excel 97-2003)  ← NLDC uses this
import openpyxl       # new .xlsx
from dotenv import load_dotenv

load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")

SHEET_NAME     = "TimeSeries"
DATA_START_ROW = 4   # 0-indexed for xlrd: row index 4 = 5th row = first data row (0:00)

# Column indices (0-indexed)
COL = {
    "time_slot":       0,
    "frequency_hz":    1,
    "demand_met_mw":   2,
    "nuclear_mw":      3,
    "wind_mw":         4,
    "solar_mw":        5,
    "hydro_mw":        6,
    "gas_mw":          7,
    "thermal_mw":      8,
    "others_mw":       9,
    "net_demand_mw":   10,
    "total_gen_mw":    11,
    "net_exchange_mw": 12,
}


# ── date helpers ──────────────────────────────────────────────────────────────

def parse_date_str(value) -> date | None:
    if value is None:
        return None
    s = str(value).strip()
    for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def date_from_filename(filepath: Path) -> date | None:
    """12.03.26_NLDC_PSP_568.xls → 2026-03-12"""
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{2})", filepath.stem)
    if m:
        day, month, yr = m.groups()
        return datetime.strptime(f"{day}.{month}.20{yr}", "%d.%m.%Y").date()
    return None


def xlrd_time_to_str(cell_val) -> str | None:
    """
    xlrd returns times as floats (fraction of a day).
    e.g. 0.0 = 00:00, 0.5 = 12:00, 0.010416... = 00:15
    """
    if not isinstance(cell_val, float):
        s = str(cell_val).strip()
        return s if re.match(r"^\d{1,2}:\d{2}", s) else None
    total_minutes = round(cell_val * 24 * 60)
    h, m = divmod(total_minutes, 60)
    return f"{h:02d}:{m:02d}"


# ── parser ────────────────────────────────────────────────────────────────────

def parse_xls(filepath: Path) -> tuple[date | None, list[dict]]:
    """Handle old .xls format using xlrd."""
    wb = xlrd.open_workbook(str(filepath))

    if SHEET_NAME not in wb.sheet_names():
        print(f"  ERROR: No sheet '{SHEET_NAME}'. Found: {wb.sheet_names()}")
        return None, []

    ws = wb.sheet_by_name(SHEET_NAME)

    # Date is in cell row=0, col=12 (M1 in Excel = index [0][12])
    date_val = ws.cell_value(0, 12)
    reading_date = parse_date_str(date_val) or date_from_filename(filepath)
    if not reading_date:
        print(f"  ERROR: Could not read date. Cell M1 = '{date_val}'")
        return None, []

    rows = []
    for row_idx in range(DATA_START_ROW, ws.nrows):
        raw = ws.row_values(row_idx)

        # Skip if row is too short or time cell is empty
        if len(raw) < 12 or raw[COL["time_slot"]] in (None, ""):
            continue

        time_str = xlrd_time_to_str(raw[COL["time_slot"]])
        if not time_str:
            continue

        def val(col_name, cast=int):
            v = raw[COL[col_name]]
            if v in (None, ""):
                return None
            try:
                return cast(v)
            except (ValueError, TypeError):
                return None

        rows.append({
            "reading_date":   reading_date,
            "time_slot":      time_str,
            "frequency_hz":   val("frequency_hz", float),
            "demand_met_mw":  val("demand_met_mw"),
            "nuclear_mw":     val("nuclear_mw"),
            "wind_mw":        val("wind_mw"),
            "solar_mw":       val("solar_mw"),
            "hydro_mw":       val("hydro_mw"),
            "gas_mw":         val("gas_mw"),
            "thermal_mw":     val("thermal_mw"),
            "others_mw":      val("others_mw"),
            "net_demand_mw":  val("net_demand_mw"),
            "total_gen_mw":   val("total_gen_mw"),
            "net_exchange_mw":val("net_exchange_mw"),
        })

    return reading_date, rows


def parse_xlsx(filepath: Path) -> tuple[date | None, list[dict]]:
    """Handle new .xlsx format using openpyxl."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"  ERROR: No sheet '{SHEET_NAME}'. Found: {wb.sheetnames}")
        return None, []

    ws = wb[SHEET_NAME]
    reading_date = parse_date_str(ws["M1"].value) or date_from_filename(filepath)
    if not reading_date:
        print(f"  ERROR: Could not read date. Cell M1 = '{ws['M1'].value}'")
        return None, []

    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        time_val = row[0]
        if time_val is None:
            continue
        time_str = time_val.strftime("%H:%M") if hasattr(time_val, "strftime") else str(time_val).strip()
        if not re.match(r"^\d{1,2}:\d{2}$", time_str):
            continue

        def v(idx, cast=int):
            val = row[idx] if idx < len(row) else None
            try:
                return cast(val) if val is not None else None
            except (ValueError, TypeError):
                return None

        rows.append({
            "reading_date":    reading_date,
            "time_slot":       time_str,
            "frequency_hz":    v(1, float),
            "demand_met_mw":   v(2),
            "nuclear_mw":      v(3),
            "wind_mw":         v(4),
            "solar_mw":        v(5),
            "hydro_mw":        v(6),
            "gas_mw":          v(7),
            "thermal_mw":      v(8),
            "others_mw":       v(9),
            "net_demand_mw":   v(10),
            "total_gen_mw":    v(11),
            "net_exchange_mw": v(12),
        })

    wb.close()
    return reading_date, rows


def parse_file(filepath: Path) -> tuple[date | None, list[dict]]:
    """Route to correct parser based on file extension."""
    ext = filepath.suffix.lower()
    if ext == ".xls":
        return parse_xls(filepath)
    elif ext == ".xlsx":
        return parse_xlsx(filepath)
    else:
        print(f"  ERROR: Unknown extension '{ext}' — expected .xls or .xlsx")
        return None, []


# ── database ──────────────────────────────────────────────────────────────────

def setup_schema():
    schema_path = Path(__file__).parent / "schema.sql"
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    cur.execute(schema_path.read_text())
    conn.commit()
    cur.close()
    conn.close()
    print("✓ Schema created / already exists.")


def insert_rows(rows: list[dict]) -> tuple[int, int]:
    if not rows:
        return 0, 0
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    inserted = skipped = 0
    for row in rows:
        cur.execute(
            """
            INSERT INTO scada_readings (
                reading_date, time_slot, frequency_hz,
                demand_met_mw, nuclear_mw, wind_mw, solar_mw,
                hydro_mw, gas_mw, thermal_mw, others_mw,
                net_demand_mw, total_gen_mw, net_exchange_mw
            ) VALUES (
                %(reading_date)s, %(time_slot)s, %(frequency_hz)s,
                %(demand_met_mw)s, %(nuclear_mw)s, %(wind_mw)s, %(solar_mw)s,
                %(hydro_mw)s, %(gas_mw)s, %(thermal_mw)s, %(others_mw)s,
                %(net_demand_mw)s, %(total_gen_mw)s, %(net_exchange_mw)s
            )
            ON CONFLICT (reading_date, time_slot) DO NOTHING
            """,
            row,
        )
        inserted += cur.rowcount
        skipped  += 1 - cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return inserted, skipped


# ── orchestration ─────────────────────────────────────────────────────────────

def process_file(filepath: Path):
    print(f"\nProcessing: {filepath.name}")
    reading_date, rows = parse_file(filepath)
    if not rows:
        print("  No rows extracted — skipping.")
        return
    flag = "✓" if len(rows) == 96 else f"⚠  expected 96, got {len(rows)}"
    print(f"  Date: {reading_date}  |  Rows: {len(rows)}  {flag}")
    inserted, skipped = insert_rows(rows)
    print(f"  Inserted: {inserted}   Skipped (duplicate): {skipped}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Parse NLDC XLS files into Neon Postgres")
    ap.add_argument("--setup",  action="store_true", help="Create DB table and exit")
    group = ap.add_mutually_exclusive_group()
    group.add_argument("--file",   type=Path, help="Single XLS/XLSX file")
    group.add_argument("--folder", type=Path, help="Folder of XLS/XLSX files")
    args = ap.parse_args()

    if args.setup:
        setup_schema()
        return

    if not args.file and not args.folder:
        ap.print_help()
        sys.exit(1)

    if args.file:
        process_file(args.file)
    elif args.folder:
        files = sorted(args.folder.glob("*.xls")) + sorted(args.folder.glob("*.xlsx"))
        if not files:
            print(f"No XLS files found in {args.folder}/")
            sys.exit(1)
        print(f"Found {len(files)} file(s)")
        for f in files:
            process_file(f)

    print("\nDone.")


if __name__ == "__main__":
    main()